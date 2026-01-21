import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import re
import numpy as np

# =========================================================
# Census Generator Tool
# - Reads ADP Data + Mapping Input
# - Reads Template .xlsm
# - Maps data and populates Template
# =========================================================

APP_TITLE = "Uzio Census Generator"

st.set_page_config(page_title=APP_TITLE, layout="centered")

st.title(APP_TITLE)
st.markdown("Upload your Input Data (ADP + Mapping) and the Census Template to generate the populated file.")

# File Uploaders
input_file = st.file_uploader("Upload Input File (Excel)", type=["xlsx", "xlsm"], key="input")
template_file = st.file_uploader("Upload Census Template (Excel .xlsm)", type=["xlsm", "xlsx"], key="template")

def norm_colname(c):
    if c is None: return ""
    return str(c).strip()

def norm_blank(x):
    if pd.isna(x): return ""
    val = str(x).strip()
    return val


# ------------------ BUSINESS LOGIC ------------------

JOB_TITLE_MAPPINGS = {
    "admin": "administrator",
    "management": "manager",
    "dsp owner": "owner"
}

TERM_REASON_MAPPINGS = {
    "no-show (never started employment)": "No-show (Never started employment)",
    "personal": "Voluntary Termination of Employment",
    "quit without notice": "Voluntary Termination of Employment",
    "attendance": "Involuntary Termination of Employment",
    "no reason given": "Involuntary Termination of Employment",
    "performance": "Involuntary Termination of Employment",
    "advancement (better job with higher pay)": "Voluntary Termination of Employment",
    "misconduct": "Involuntary Termination of Employment",
    "mutual agreement": "Voluntary Termination of Employment",
    "abandoned job": "Involuntary Termination of Employment",
    "deceased": "Death",
    "retirement": "Retirement",
    "permanent disability": "Permanent Disability",
    "transfer": "Transfer",
    # Mappings for lowercased inputs to standard output for direct hits
    "involuntary termination of employment": "Involuntary Termination of Employment",
    "voluntary termination of employment": "Voluntary Termination of Employment",
    # Add other direct keys just in case
    "death": "Death",
    "other": "Other"
}

STATE_MAP = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR", "california": "CA",
    "colorado": "CO", "connecticut": "CT", "delaware": "DE", "florida": "FL", "georgia": "GA",
    "hawaii": "HI", "idaho": "ID", "illinois": "IL", "indiana": "IN", "iowa": "IA",
    "kansas": "KS", "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN", "mississippi": "MS",
    "missouri": "MO", "montana": "MT", "nebraska": "NE", "nevada": "NV", "new hampshire": "NH",
    "new jersey": "NJ", "new mexico": "NM", "new york": "NY", "north carolina": "NC",
    "north dakota": "ND", "ohio": "OH", "oklahoma": "OK", "oregon": "OR", "pennsylvania": "PA",
    "rhode island": "RI", "south carolina": "SC", "south dakota": "SD", "tennessee": "TN",
    "texas": "TX", "utah": "UT", "vermont": "VT", "virginia": "VA", "washington": "WA",
    "west virginia": "WV", "wisconsin": "WI", "wyoming": "WY",
    "district of columbia": "DC"
}

def norm_job_title(x):
    if not x: return ""
    s = str(x).strip().replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s).lower()
    return JOB_TITLE_MAPPINGS.get(s, s)

def get_mapped_term_reason(x):
    if not x: return "Other"
    s = str(x).strip().replace("\u00A0", " ")
    if not s or s.lower() == "nan": return "Other"
    
    s_lower = re.sub(r"\s+", " ", s).lower()
    
    # Direct look up
    if s_lower in TERM_REASON_MAPPINGS:
        return TERM_REASON_MAPPINGS[s_lower]
    
    # If not found in explicit map, default to Other
    return "Other"

def get_state_abbr(x):
    if not x: return ""
    s = str(x).strip()
    if not s or s.lower() == "nan": return ""
    
    if len(s) <= 2:
        return s.upper()
    
    s_lower = s.lower()
    return STATE_MAP.get(s_lower, s)

# Copy of deduplication logic to keep this tool independent
def deduplicate_adp(df: pd.DataFrame, key_col: str) -> pd.DataFrame:
    # Identify special columns (normalized)
    col_map = {c: c.lower() for c in df.columns}
    
    status_col = next((c for c, l in col_map.items() if "position status" in l), None)
    term_date_col = next((c for c, l in col_map.items() if "termination date" in l), None)
    start_date_col = next((c for c, l in col_map.items() if "position start date" in l), None)
    loc_desc_col = next((c for c, l in col_map.items() if "work location description" in l), None)
    license_id_col = next((c for c, l in col_map.items() if "license/certification id" in l), None)
    
    # If we can't find status col, fallback to basic drop_duplicates
    if not status_col:
        return df.drop_duplicates(subset=[key_col], keep="first")
        
    def pick_best(group):
        if len(group) <= 1:
            return group.iloc[[0]]
        
        # Helper to parse date for sorting
        def get_date_val(row, col):
            if not col or pd.isna(row[col]):
                return pd.Timestamp.min
            val = str(row[col]).strip()
            if not val:
                return pd.Timestamp.min
            try:
                return pd.to_datetime(val)
            except:
                return pd.Timestamp.min

        if isinstance(group, pd.Series):
                return group.to_frame().T

        group = group.copy()
        group['__norm_status'] = group[status_col].astype(str).str.lower().str.strip()
        
        # Add license check
        if license_id_col:
            group['__has_license'] = group[license_id_col].apply(lambda x: 1 if norm_blank(x) != "" else 0)
        else:
            group['__has_license'] = 0

        actives = group[group['__norm_status'] == 'active']
        terms = group[group['__norm_status'] == 'terminated']
        others = group[(group['__norm_status'] != 'active') & (group['__norm_status'] != 'terminated')]
        
        # Logic 1: If Actives exist, prioritize them
        if not actives.empty:
            actives['__sort_date'] = actives.apply(lambda r: get_date_val(r, start_date_col), axis=1)

            if loc_desc_col:
                actives['__has_loc'] = actives[loc_desc_col].apply(lambda x: 1 if norm_blank(x) != "" else 0)
                best_active = actives.sort_values(by=['__has_loc', '__has_license', '__sort_date'], ascending=[False, False, False]).iloc[[0]]
            else:
                best_active = actives.sort_values(by=['__has_license', '__sort_date'], ascending=[False, False]).iloc[[0]]
            
            return best_active

        # Logic 2: Terminated
        if not terms.empty:
            terms['__sort_date'] = pd.Timestamp.min
            
            use_start_date = False
            if term_date_col:
                terms['__term_dt_val'] = terms[term_date_col].apply(norm_blank)
                has_blank = (terms['__term_dt_val'] == "").any()
                has_val = (terms['__term_dt_val'] != "").any()
                
                if has_blank and has_val:
                    use_start_date = True
            else:
                use_start_date = True

            if use_start_date:
                    terms['__sort_date'] = terms.apply(lambda r: get_date_val(r, start_date_col), axis=1)
            elif term_date_col:
                    terms['__sort_date'] = terms.apply(lambda r: get_date_val(r, term_date_col), axis=1)
            
            return terms.sort_values(by=['__has_license', '__sort_date'], ascending=[False, False]).iloc[[0]]

        # Fallback (Others)
        if not others.empty:
                others['__sort_date'] = others.apply(lambda r: get_date_val(r, start_date_col), axis=1)
                return others.sort_values(by=['__has_license', '__sort_date'], ascending=[False, False]).iloc[[0]]

        return group.iloc[[0]]

    # Apply grouping
    deduped = df.groupby(key_col, group_keys=False).apply(pick_best)
    
    # Cleanup temp columns
    cols_to_drop = [c for c in ['__norm_status', '__has_loc', '__sort_date', '__term_dt_val', '__has_license'] if c in deduped.columns]
    if cols_to_drop:
        deduped = deduped.drop(columns=cols_to_drop)
        
    return deduped

def process_files(input_io, template_io):
    # 1. Read Input Data
    try:
        xls_in = pd.ExcelFile(input_io, engine="openpyxl")
        # Assuming typical sheet names or indices: 0=ADP, 1=Mapping? 
        # User said "tab 1 he uploads the ADP data and in tab 2 he uploads the mapping"
        # Tab 1 at index 0, Tab 2 at index 1
        sheet_map = {s.lower(): s for s in xls_in.sheet_names}
        
        # Determine ADP Sheet
        adp_sheet = sheet_map.get("adp data")
        if not adp_sheet:
            adp_sheet = sheet_map.get("adp")
        if not adp_sheet:
            adp_sheet = 0 # Fallback
            
        # Determine Mapping Sheet
        map_sheet_name = sheet_map.get("mapping sheet")
        if not map_sheet_name:
            map_sheet_name = sheet_map.get("mapping")
        if not map_sheet_name:
            if len(xls_in.sheet_names) > 1:
                map_sheet_name = 1
            else:
                map_sheet_name = 1

        if len(xls_in.sheet_names) < 2 and (adp_sheet == 0 or map_sheet_name == 1) and len(sheet_map) < 2:
            st.error("Input file must have at least 2 sheets.")
            return None
             
        adp_df = pd.read_excel(xls_in, sheet_name=adp_sheet, dtype=object)
        map_sheet = pd.read_excel(xls_in, sheet_name=map_sheet_name, header=None, dtype=object)
        
        # Find Header Row in Mapping
        header_row_idx = None
        for i, row in map_sheet.iterrows():
            row_str = row.astype(str).str.strip().str.lower().tolist()
            if "uzio coloumn" in row_str or "uzio column" in row_str:
                header_row_idx = i
                break
        
        if header_row_idx is None:
            st.error("Could not find 'Uzio Coloumn' header in Mapping sheet.")
            return None
            
        # Reload with correct header
        map_sheet.columns = map_sheet.iloc[header_row_idx].astype(str)
        map_df = map_sheet.iloc[header_row_idx+1:].reset_index(drop=True)
        
    except Exception as e:
        st.error(f"Error reading input file: {e}")
        return None

    # 2. Normalize Headers
    adp_df.columns = [norm_colname(c) for c in adp_df.columns]
    map_df.columns = [norm_colname(c) for c in map_df.columns]

    
    # 3. Process Mapping
    if "Uzio Coloumn" not in map_df.columns or "ADP Coloumn" not in map_df.columns:
        st.error("Mapping sheet must contain 'Uzio Coloumn' and 'ADP Coloumn'.")
        return None
        
    # Filter valid mappings
    mapping_valid = map_df.dropna(subset=["Uzio Coloumn", "ADP Coloumn"])
    mapping_valid = mapping_valid[(mapping_valid["Uzio Coloumn"] != "") & (mapping_valid["ADP Coloumn"] != "")]
    
    # Create Map Dictionary: {TemplateHeader: ADPHeader}
    # Note: Using "Uzio Coloumn" as Template Header
    col_map = dict(zip(mapping_valid["Uzio Coloumn"].astype(str).str.strip(), mapping_valid["ADP Coloumn"].astype(str).str.strip()))
    
    # 4. Find Key for Deduplication
    key_row = mapping_valid[mapping_valid["Uzio Coloumn"].str.contains("Associate ID|Employee ID", case=False, na=False)]
    if len(key_row) > 0:
        ADP_KEY = key_row.iloc[0]["ADP Coloumn"]
        if ADP_KEY in adp_df.columns:
            st.info(f"Deduplicating ADP data using key: {ADP_KEY}...")
            # Normalize key column
            adp_df[ADP_KEY] = adp_df[ADP_KEY].astype(str).str.strip()
            adp_df = deduplicate_adp(adp_df, ADP_KEY)
            st.success(f"Deduplication complete. Rows: {len(adp_df)}")
        else:
            st.warning(f"Key column '{ADP_KEY}' not found in ADP data. Skipping deduplication.")
    else:
        st.warning("Could not identify EmployeeID/AssociateID mapping. Skipping deduplication.")

    # 5. Load Template
    try:
        # keep_vba=True is crucial for .xlsm
        wb = openpyxl.load_workbook(template_io, keep_vba=True)
        # Assuming sheet of interest matches the sample: "Employee Census" (Index 1 usually, or name based)
        # User sample had 'Instructions' at 0, 'Employee Census' at 1. 
        # Let's try to find 'Employee Census' or 'Census' or fall back to index 1
        target_sheet_name = None
        for sn in wb.sheetnames:
            if "census" in sn.lower():
                target_sheet_name = sn
                break
        
        if not target_sheet_name:
            if len(wb.sheetnames) > 1:
                target_sheet_name = wb.sheetnames[1]
            else:
                target_sheet_name = wb.sheetnames[0]
                
        ws = wb[target_sheet_name]
        st.info(f"Targeting Template Sheet: {target_sheet_name}")

        # 6. Identify Header Row & Build Column Maps
        # Heuristic: Find row with most matches to our Mapping Keys
        header_row_idx = None
        header_map = {} # {ColumnIndex (1-based): ADPHeaderName}
        
        max_matches = 0
        best_row = 1
        
        # Scan first 10 rows
        for r in range(1, 11):
            row_values = []
            for cell in ws[r]:
                row_values.append(str(cell.value).strip() if cell.value else "")
            
            matches = sum(1 for v in row_values if v in col_map)
            if matches > max_matches:
                max_matches = matches
                best_row = r
                
        if max_matches == 0:
            st.error("Could not locate header row in Template matching your Mapping keys. Please check mapping vs template headers.")
            return None
            
        header_row_idx = best_row
        
        # Build FULL Template Header Map (Index -> Name) and Mapped Columns (Index -> ADP Header)
        all_template_headers = {} # {ColIdx: HeaderName}
        
        for cell in ws[header_row_idx]:
            val = str(cell.value).strip() if cell.value else ""
            if not val: continue
            
            all_template_headers[cell.column] = val
            
            if val in col_map:
                header_map[cell.column] = col_map[val] # Map ColIndex -> ADPHeader

        st.info(f"Identified Header Row at {header_row_idx}. Mapped {len(header_map)} columns linked to ADP data.")

        # 7. Write Data
        # Start writing at header_row + 1
        start_row = header_row_idx + 1
        
        # Convert ADP dataframe to list of dicts for easy access
        adp_records = adp_df.to_dict('records')

        # --- SPECIAL COLUMN IDENTIFICATION ---
        # Look through ALL template headers to find special functional columns
        idx_salary = []
        idx_hours = []
        idx_hourly_rate = [] 
        idx_pay_type = []
        idx_term_reason = []
        idx_job_title = []
        idx_state = []
        idx_union = [] 
        idx_flsa = []
        idx_dates = [] # New: Date columns
        
        def is_header(h_name, keywords):
            if not h_name: return False
            h = str(h_name).lower().strip()
            return any(k in h for k in keywords)

        for c_idx, h_name in all_template_headers.items():
            # Salary
            if is_header(h_name, ["annual salary"]):
                idx_salary.append(c_idx)
            # Hourly Pay Rate
            elif is_header(h_name, ["hourly pay", "hourly rate"]):
                 idx_hourly_rate.append(c_idx)
            # Hours
            elif is_header(h_name, ["standard hours", "working hours"]):
                idx_hours.append(c_idx)
            # Pay Type
            elif is_header(h_name, ["pay type", "employment type"]): 
                idx_pay_type.append(c_idx)
            # Term Reason
            elif is_header(h_name, ["termination reason"]):
                idx_term_reason.append(c_idx)
            # Job Title
            elif is_header(h_name, ["job title"]):
                idx_job_title.append(c_idx)
            # State
            elif is_header(h_name, ["state", "work state"]): 
                 if "tax" not in h_name.lower():
                     idx_state.append(c_idx)
            # Union
            elif is_header(h_name, ["union classification", "union"]):
                idx_union.append(c_idx)
            # FLSA
            elif is_header(h_name, ["flsa classification", "flsa"]):
                idx_flsa.append(c_idx)
            # Dates
            elif is_header(h_name, ["date", "dob", "birth", "expire"]): 
                idx_dates.append(c_idx)

        # Merge all columns we need to write to: Mapped Columns + Special Static Columns
        target_columns = set(header_map.keys())
        target_columns.update(idx_union)
        target_columns.update(idx_flsa)
        
        processed_count = 0
        for i, record in enumerate(adp_records):
            current_row = start_row + i
            
            # --- ROW LEVEL CONTEXT ---
            # Get Pay Type Value first
            row_pay_type_source = ""
            if idx_pay_type:
                # Use the first Pay Type column found
                pt_col_idx = idx_pay_type[0]
                # Is this column mapped from ADP?
                adp_header_for_pt = header_map.get(pt_col_idx)
                if adp_header_for_pt:
                   row_pay_type_source = str(record.get(adp_header_for_pt, "")).lower()

            # Normalize Pay Type
            norm_pay_type_val = ""
            is_hourly = False
            is_salary = False
            
            if "hour" in row_pay_type_source:
                norm_pay_type_val = "Hourly"
                is_hourly = True
            elif "sal" in row_pay_type_source or "exempt" in row_pay_type_source:
                norm_pay_type_val = "Salaried"
                is_salary = True
            else:
                 if row_pay_type_source:
                     norm_pay_type_val = row_pay_type_source.title()
                 else:
                     norm_pay_type_val = "" 
            
            if norm_pay_type_val.lower() == "salary":
                norm_pay_type_val = "Salaried"

            # Iterate over all target columns
            for col_idx in target_columns:
                # Default value from ADP if mapped
                adp_header = header_map.get(col_idx)
                val = ""
                if adp_header:
                    val = record.get(adp_header, "")
                    if pd.isna(val): val = ""

                # --- APPLY TRANSFORMATIONS ---

                # 0. Pay Type Normalization
                if col_idx in idx_pay_type:
                    val = norm_pay_type_val if norm_pay_type_val else val
                
                # 1. Salary / Hours / Hourly Rate Clearing
                elif col_idx in idx_salary:
                    if is_hourly: val = "" 
                elif col_idx in idx_hours:
                    if is_salary: val = ""
                elif col_idx in idx_hourly_rate:
                    if is_salary: val = ""
                        
                # 2. Termination Reason
                elif col_idx in idx_term_reason:
                    val = get_mapped_term_reason(val)
                    
                # 3. Job Title
                elif col_idx in idx_job_title:
                    val = norm_job_title(val)
                    
                # 4. State
                elif col_idx in idx_state:
                    val = get_state_abbr(val)
                
                # 5. Union Classification (Always Non-Union)
                elif col_idx in idx_union:
                    val = "Non-Union"
                    
                # 6. FLSA Classification
                elif col_idx in idx_flsa:
                    if is_hourly:
                        val = "Non-Exempt"
                    elif is_salary:
                        val = "Exempt"
                
                # 7. Date Formatting
                # Determine if we should attempt formatting
                if col_idx in idx_dates:
                     if val and str(val).strip():
                         try:
                             # Try to parse and format as YYYY-MM-DD
                             dt = pd.to_datetime(val)
                             val = dt.strftime("%Y-%m-%d")
                         except:
                             pass # Keep original if parse fails

                ws.cell(row=current_row, column=col_idx, value=val)
            processed_count += 1
            
        st.success(f"Populated {processed_count} rows.")

        # 8. Save
        out_buffer = io.BytesIO()
        wb.save(out_buffer)
        out_buffer.seek(0)
        return out_buffer

    except Exception as e:
        st.error(f"Error processing template: {e}")
        return None

if input_file and template_file:
    if st.button("Generate Census"):
        with st.spinner("Processing..."):
            result_xlsm = process_files(input_file, template_file)
            
            if result_xlsm:
                st.download_button(
                    label="Download Populated Census (.xlsm)",
                    data=result_xlsm,
                    file_name="Populated_Census.xlsm",
                    mime="application/vnd.ms-excel.sheet.macroEnabled.12"
                )
